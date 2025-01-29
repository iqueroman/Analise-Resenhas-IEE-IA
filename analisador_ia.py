import logging
from typing import Dict, Any, List, Tuple
from pathlib import Path
import pandas as pd
from detector_gpt_zero import GPTZeroDetector
from detector_zero_gpt import ZeroGPTDetector
import openpyxl
from time import sleep
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt

class AnalisadorIA:
    def __init__(self, gpt_zero_key: str, zero_gpt_key: str):
        self.logger = logging.getLogger('detector_ia')
        self.gpt_zero = GPTZeroDetector(gpt_zero_key)
        self.zero_gpt = ZeroGPTDetector(zero_gpt_key)
    
    def analisar_resumos(self, resumos: List[Tuple[str, str]]) -> List[Dict]:
        """
        Analisa uma lista de resumos e retorna resultados para cada um,
        processando em lotes de 40 para respeitar limites da API
        """
        resultados = []
        tamanho_lote = 40
        
        for i in range(0, len(resumos), tamanho_lote):
            lote_atual = resumos[i:i + tamanho_lote]
            self.logger.info(f"Processando lote {(i//tamanho_lote)+1} ({len(lote_atual)} textos)")
            
            for nome_livro, texto in lote_atual:
                self.logger.info(f"Analisando resumo: {nome_livro}")
                resultado = {
                    'Livro': nome_livro,
                    'Texto_Normalizado': texto
                }
                
                # Análise GPTZero
                try:
                    gpt_zero_result = self.gpt_zero.analisar_texto(texto)
                    resultado.update({
                        # Metadados
                        'GPTZero_Versao': gpt_zero_result['version'],
                        'GPTZero_ScanID': gpt_zero_result['scan_id'],
                        
                        # Probabilidades principais
                        'GPTZero_Prob_Media_IA': gpt_zero_result['documento']['prob_media_ia'],
                        'GPTZero_Prob_IA': gpt_zero_result['documento']['prob_classes']['ai'],
                        'GPTZero_Prob_Humano': gpt_zero_result['documento']['prob_classes']['human'],
                        'GPTZero_Prob_Misto': gpt_zero_result['documento']['prob_classes']['mixed'],
                        
                        # Confiança e classificação
                        'GPTZero_Categoria_Confianca': gpt_zero_result['documento']['categoria_confianca'],
                        'GPTZero_Pontuacao_Confianca': gpt_zero_result['documento']['pontuacao_confianca'],
                        'GPTZero_Classe_Prevista': gpt_zero_result['documento']['classe_prevista'],
                        'GPTZero_Classificacao': gpt_zero_result['documento']['classificacao_documento'],
                        
                        # Mensagem
                        'GPTZero_Mensagem': gpt_zero_result['documento']['mensagem_resultado'],
                        
                        # Sentenças destacadas
                        'GPTZero_Sentencas_Destacadas': '; '.join([
                            s['texto'] for s in gpt_zero_result['sentencas'] 
                            if s['destacar_ia']
                        ])
                    })
                except Exception as e:
                    self.logger.error(f"Erro na análise GPTZero de {nome_livro}: {str(e)}")
                    resultado.update({
                        'GPTZero_Versao': None,
                        'GPTZero_ScanID': None,
                        'GPTZero_Prob_Media_IA': -1,
                        'GPTZero_Prob_IA': -1,
                        'GPTZero_Prob_Humano': -1,
                        'GPTZero_Prob_Misto': -1,
                        'GPTZero_Categoria_Confianca': None,
                        'GPTZero_Pontuacao_Confianca': -1,
                        'GPTZero_Classe_Prevista': None,
                        'GPTZero_Classificacao': None,
                        'GPTZero_Mensagem': None,
                        'GPTZero_Sentencas_Destacadas': None
                    })
                
                # Análise ZeroGPT
                try:
                    zero_gpt_result = self.zero_gpt.analisar_texto(texto)
                    
                    # Formata as sentenças IA para exibição no Excel
                    sentencas_ia = zero_gpt_result['sentencas_ia']
                    if sentencas_ia:
                        # Remove caracteres problemáticos e formata cada sentença
                        sentencas_formatadas = [
                            f"{i+1}. {str(sentenca).strip().replace('[', '(').replace(']', ')')}"
                            for i, sentenca in enumerate(sentencas_ia)
                        ]
                        # Junta as sentenças com quebra de linha
                        texto_sentencas = "\n".join(sentencas_formatadas)
                    else:
                        texto_sentencas = "Nenhuma sentença identificada como IA"
                    
                    resultado.update({
                        'ZeroGPT_Sucesso': zero_gpt_result['success'],
                        'ZeroGPT_Total_Palavras': zero_gpt_result['total_palavras'],
                        'ZeroGPT_Palavras_IA': zero_gpt_result['palavras_ia'],
                        'ZeroGPT_Porcentagem_IA': zero_gpt_result['porcentagem_ia'],
                        'ZeroGPT_Sentencas_IA': texto_sentencas,  # Sentenças formatadas
                        'ZeroGPT_Feedback': zero_gpt_result['feedback'],
                        'ZeroGPT_Mensagem': zero_gpt_result['mensagem']
                    })
                except Exception as e:
                    self.logger.error(f"Erro na análise ZeroGPT de {nome_livro}: {str(e)}")
                    resultado.update({
                        'ZeroGPT_Sucesso': None,
                        'ZeroGPT_Total_Palavras': None,
                        'ZeroGPT_Palavras_IA': None,
                        'ZeroGPT_Porcentagem_IA': None,
                        'ZeroGPT_Sentencas_IA': None,
                        'ZeroGPT_Feedback': None,
                        'ZeroGPT_Mensagem': None
                    })
                
                # Adiciona delay entre chamadas
                sleep(1)
                resultados.append(resultado)
            
            # Delay entre lotes
            if i + tamanho_lote < len(resumos):
                self.logger.info("Aguardando 5 segundos antes do próximo lote...")
                sleep(5)
        
        return resultados

def gerar_relatorio_completo(resultados_participante: List[Dict], pasta_base: Path, participante: str):
    logger = logging.getLogger('detector_ia')
    pasta_relatórios = pasta_base / "Relatórios"
    pasta_relatórios.mkdir(exist_ok=True)
    
    try:
        # Limpa os nomes dos livros removendo caracteres especiais
        for resultado in resultados_participante:
            resultado['Livro'] = resultado['Livro'].replace('[', '').replace(']', '')
        
        # Cria o DataFrame
        df = pd.DataFrame(resultados_participante)
        
        # Renomeia as colunas
        df = df.rename(columns={
            'Livro': 'Livro/curso',
            'Texto_Normalizado': 'Resenha'
        })
        
        excel_file = pasta_relatórios / f"relatório_{participante}.xlsx"
        
        # Configura a largura das colunas e adiciona comentários explicativos
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Análises')
            worksheet = writer.sheets['Análises']
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")  # Fonte branca para cabeçalhos
            gptzero_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul para GPTZero
            zerogpt_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Verde para ZeroGPT
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplica estilos aos cabeçalhos
            for idx, column in enumerate(df.columns):
                col_letter = get_column_letter(idx + 1)
                cell = worksheet[f"{col_letter}1"]
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)  # Fonte preta em negrito por padrão
                
                # Aplica cor branca e fundo colorido apenas para colunas das APIs
                if column.startswith('GPTZero'):
                    cell.font = header_font  # Fonte branca
                    cell.fill = gptzero_fill
                elif column.startswith('ZeroGPT'):
                    cell.font = header_font  # Fonte branca
                    cell.fill = zerogpt_fill
            
            # Definição das cores base para o degradê
            VERDE = "63BE7B"  # Verde mais suave
            AMARELO = "FFEB84"
            VERMELHO = "F8696B"  # Vermelho mais suave

            # Lista de colunas que usam o degradê padrão (0=verde, 1=vermelho)
            prob_columns_ia = [
                'GPTZero_Prob_Media_IA',  # 0=humano/verde, 1=IA/vermelho
                'GPTZero_Prob_IA',        # 0=humano/verde, 1=IA/vermelho
                'GPTZero_Prob_Misto',     # 0=humano/verde, 1=IA/vermelho
                'ZeroGPT_Porcentagem_IA'  # 0=humano/verde, 1=IA/vermelho
            ]

            # Lista de colunas que usam o degradê invertido (0=vermelho, 1=verde)
            prob_columns_humano = [
                'GPTZero_Prob_Humano',    # 0=IA/vermelho, 1=humano/verde
                'GPTZero_Pontuacao_Confianca'  # 0=baixa confiança/vermelho, 1=alta confiança/verde
            ]

            # Formatação para campos categóricos
            categoria_rules = {
                'GPTZero_Categoria_Confianca': {
                    'high': PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid"),
                    'medium': PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid"),
                    'low': PatternFill(start_color=VERMELHO, end_color=VERMELHO, fill_type="solid")
                },
                'GPTZero_Classe_Prevista': {
                    'human': PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid"),
                    'mixed': PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid"),
                    'ai': PatternFill(start_color=VERMELHO, end_color=VERMELHO, fill_type="solid")
                },
                'GPTZero_Classificacao': {
                    'human_only': PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid"),
                    'mixed': PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid"),
                    'ai_only': PatternFill(start_color=VERMELHO, end_color=VERMELHO, fill_type="solid")
                },
                'ZeroGPT_Feedback': {
                    'your text is human written': PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid"),
                    'your text is most likely': PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid"),
                    'your text is ai/gpt generated': PatternFill(start_color=VERMELHO, end_color=VERMELHO, fill_type="solid")
                }
            }

            # Aplica formatação para campos categóricos
            for col, regras in categoria_rules.items():
                if col in df.columns:
                    col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                    for row_idx in range(2, len(df) + 2):  # Começa da linha 2 (após cabeçalho)
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        valor = str(cell.value).lower() if cell.value else ''
                        # Verifica se o valor começa com alguma das chaves (para match parcial)
                        for key, fill in regras.items():
                            if valor.startswith(key):
                                cell.fill = fill
                                break
            
            # Cria regra de formatação com 100 níveis para IA (0=verde, 1=vermelho)
            color_scale_ia = ColorScaleRule(
                start_type='num', start_value=0, start_color=VERDE,
                mid_type='num', mid_value=0.5, mid_color=AMARELO,
                end_type='num', end_value=1, end_color=VERMELHO
            )

            # Cria regra de formatação com 100 níveis para humano (0=vermelho, 1=verde)
            color_scale_humano = ColorScaleRule(
                start_type='num', start_value=0, start_color=VERMELHO,
                mid_type='num', mid_value=0.5, mid_color=AMARELO,
                end_type='num', end_value=1, end_color=VERDE
            )

            # Aplica formatação para indicadores de IA
            for col in prob_columns_ia:
                if col in df.columns:
                    col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                    worksheet.conditional_formatting.add(
                        f"{col_letter}2:{col_letter}{len(df)+1}",
                        color_scale_ia
                    )

            # Aplica formatação para indicadores humanos
            for col in prob_columns_humano:
                if col in df.columns:
                    col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                    worksheet.conditional_formatting.add(
                        f"{col_letter}2:{col_letter}{len(df)+1}",
                        color_scale_humano
                    )
            
            # Ajusta largura das colunas
            for idx, column in enumerate(df.columns):
                col_letter = get_column_letter(idx + 1)
                if column == 'Livro':
                    width = 30
                elif column == 'Resenha':
                    width = 50
                elif column == 'ZeroGPT_Sentencas_IA':  # Coluna específica para sentenças
                    width = 60  # Largura maior para acomodar as sentenças
                elif 'Mensagem' in column or 'Feedback' in column:
                    width = 30
                else:
                    width = 15
                worksheet.column_dimensions[col_letter].width = width
            
            # Ajusta altura das linhas para serem pequenas
            for row_idx in range(2, len(df) + 2):  # Começa da linha 2 (após cabeçalho)
                worksheet.row_dimensions[row_idx].height = 15  # Altura fixa pequena
            
            # Congela o painel para manter cabeçalhos visíveis
            worksheet.freeze_panes = 'A2'
            
            # Alinhamento para todas as células
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical='center')
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin')
                    )
            
            # Adiciona comentários explicativos
            explicacoes = {
                # GPTZero - Metadados
                'GPTZero_Versao': 'Versão do detector GPTZero usado na análise',
                'GPTZero_ScanID': 'Identificador único do scan. Um scan pode ter múltiplos documentos',
                
                # GPTZero - Probabilidades
                'GPTZero_Prob_Media_IA': 'Média das probabilidades de cada sentença ser IA (0-1). Quanto maior, mais provável ser IA',
                'GPTZero_Prob_IA': 'Probabilidade do texto ser inteiramente IA (0-1). Use junto com Categoria_Confianca',
                'GPTZero_Prob_Humano': 'Probabilidade do texto ser inteiramente humano (0-1)',
                'GPTZero_Prob_Misto': 'Probabilidade do texto ser uma mistura de IA e humano (0-1)',
                
                # GPTZero - Confiança
                'GPTZero_Categoria_Confianca': '"high" (<1% erro), "medium" (confiança moderada), "low" (baixa confiança)',
                'GPTZero_Pontuacao_Confianca': 'Score normalizado de confiança (uso interno)',
                
                # GPTZero - Classificação
                'GPTZero_Classe_Prevista': '"human" (só humano), "ai" (só IA), "mixed" (mistura)',
                'GPTZero_Classificacao': '"HUMAN_ONLY" (predominante humano), "MIXED" (misto/fraca IA), "AI_ONLY" (todo IA)',
                'GPTZero_Mensagem': 'Ex: "highly confident text is written by AI"',
                
                # GPTZero - Sentenças
                'GPTZero_Sentencas_Destacadas': 'Sentenças identificadas como prováveis de serem IA',
                
                # ZeroGPT
                'ZeroGPT_Sucesso': 'Status da análise: true = sucesso, false = falha',
                'ZeroGPT_Total_Palavras': 'Contagem total de palavras no texto',
                'ZeroGPT_Palavras_IA': 'Número de palavras identificadas como IA',
                'ZeroGPT_Porcentagem_IA': 'Porcentagem (0-100%) do texto identificada como IA. 0% = humano, 100% = IA',
                'ZeroGPT_Sentencas_IA': 'Lista de sentenças identificadas com maior probabilidade de serem geradas por IA',
                'ZeroGPT_Feedback': 'Análise detalhada do texto',
                'ZeroGPT_Mensagem': 'Status e resultado geral da operação'
            }
            
            # Adiciona comentários nas células
            for idx, (coluna, explicacao) in enumerate(explicacoes.items()):
                if coluna in df.columns:  # Só adiciona se a coluna existir
                    col_letter = get_column_letter(df.columns.get_loc(coluna) + 1)
                    cell = worksheet[f"{col_letter}1"]
                    cell.comment = openpyxl.comments.Comment(
                        explicacao,
                        'Detector IA'
                    )
            
            # Alinhamento especial para a coluna de sentenças
            if 'ZeroGPT_Sentencas_IA' in df.columns:
                col_letter = get_column_letter(df.columns.get_loc('ZeroGPT_Sentencas_IA') + 1)
                for cell in worksheet[col_letter]:
                    if cell.row > 1:  # Pula o cabeçalho
                        cell.alignment = Alignment(
                            vertical='top',
                            wrap_text=True,
                            shrink_to_fit=True  # Adiciona shrink_to_fit para ajudar a manter o texto na linha pequena
                        )
        
        # Após gerar o Excel, cria o gráfico de dispersão individual
        df_analise = pd.DataFrame(resultados_participante)
        
        plt.figure(figsize=(10, 8))
        plt.scatter(df_analise['GPTZero_Prob_IA'], df_analise['ZeroGPT_Porcentagem_IA'])
        
        # Adiciona rótulos para cada ponto (nome do livro)
        for i, livro in enumerate(df_analise['Livro']):
            plt.annotate(livro, 
                       (df_analise['GPTZero_Prob_IA'].iloc[i], 
                        df_analise['ZeroGPT_Porcentagem_IA'].iloc[i]),
                       fontsize=8)
        
        plt.xlabel('GPTZero_Prob_IA')
        plt.ylabel('ZeroGPT_Porcentagem_IA')
        plt.title(f'Comparação entre Detectores - {participante}')
        plt.grid(True)
        
        # Define limites fixos para os eixos
        plt.xlim(0, 1)  # GPTZero vai de 0 a 1
        plt.ylim(0, 100)  # ZeroGPT vai de 0 a 100
        
        # Salva o gráfico na mesma pasta do Excel
        arquivo_grafico = pasta_relatórios / f"grafico_dispersao_{participante}.png"
        plt.savefig(arquivo_grafico, bbox_inches='tight', dpi=300)
        plt.close()
        
        logger.info(f"Gráfico de dispersão gerado para {participante}: {arquivo_grafico}")
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório para {participante}: {str(e)}", exc_info=True) 